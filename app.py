import base64
import csv
import io
import json
import os
import re
import time
import zipfile
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="DienmayXANH Product Scraper", page_icon="📦", layout="wide")

# ---------------------------------------------------------------------------
# Cookie mặc định (dienmayxanh.com) — nhúng sẵn để tool tự resolve link rút
# gọn /sp-{id} mà không cần tải cookies.json lên mỗi lần dùng (dễ quên / dễ
# nhầm file). Cookie có thể hết hạn theo thời gian — nếu resolve id bắt đầu
# lỗi 404 hàng loạt, hãy tải cookies.json mới (xuất từ trình duyệt đã đăng
# nhập dienmayxanh.com) lên ở mục "Cookie nâng cao" bên dưới để ghi đè.
# ---------------------------------------------------------------------------
DEFAULT_COOKIES = [
    {"name": "_ce.clock_data", "domain": ".dienmayxanh.com",
     "value": "-140%2C113.161.36.0%2C1%2C6967ec7261b3cbe6a91d798c6b951c60%2CChrome%2CVN%2CDirect%20Traffic%2C"},
    {"name": "_ce.s", "domain": ".dienmayxanh.com",
     "value": "v~fae3bb319ee26758400d38572a0b2badd124a3fd~vir~returning~lva~1787296884142~vpv~3~v11ls~ea633fc0-9d30-11f1-a170-29257f06f2e4~v11.cs~454123~v11.s~ea633fc0-9d30-11f1-a170-29257f06f2e4~v11.vs~fae3bb319ee26758400d38572a0b2badd124a3fd~v11.sla~1787296884159~v11.wss~1787296884160~v11.ss~1787296884163~lcw~1787296884165"},
    {"name": "_fbp", "domain": ".dienmayxanh.com", "value": "fb.1.1782793183066.926715296684095679"},
    {"name": "_ga", "domain": ".dienmayxanh.com", "value": "GA1.1.763391756.1782793167"},
    {"name": "_ga_Y7SWKJEHCE", "domain": ".dienmayxanh.com",
     "value": "GS2.1.s1787296883$o4$g0$t1787296884$j59$l0$h0"},
    {"name": "_gcl_au", "domain": ".dienmayxanh.com", "value": "1.1.1295786637.1782793167"},
    {"name": "cebs", "domain": ".dienmayxanh.com", "value": "1"},
    {"name": "cebsp_", "domain": ".dienmayxanh.com", "value": "1"},
    {"name": "DMX_Personal", "domain": ".dienmayxanh.com",
     "value": "%7B%22CustomerId%22%3A0%2C%22CustomerSex%22%3A-1%2C%22CustomerName%22%3Anull%2C%22CustomerPhone%22%3Anull%2C%22CustomerMail%22%3Anull%2C%22Lat%22%3A0.0%2C%22Lng%22%3A0.0%2C%22Address%22%3Anull%2C%22CurrentUrl%22%3Anull%2C%22ProvinceId%22%3A1027%2C%22ProvinceType%22%3Anull%2C%22ProvinceName%22%3A%22H%E1%BB%93%20Ch%C3%AD%20Minh%22%2C%22DistrictId%22%3A0%2C%22DistrictType%22%3Anull%2C%22DistrictName%22%3Anull%2C%22WardId%22%3A0%2C%22WardType%22%3Anull%2C%22WardName%22%3Anull%2C%22StoreId%22%3A0%2C%22CouponCode%22%3Anull%2C%22HasLocation%22%3Afalse%7D"},
    {"name": "mwgsp", "domain": ".dienmayxanh.com", "value": "1"},
    {"name": "__admUTMtime", "domain": ".www.dienmayxanh.com", "value": "1787293383"},
    {"name": "__iid", "domain": ".www.dienmayxanh.com", "value": ""},
    {"name": "__su", "domain": ".www.dienmayxanh.com", "value": "0"},
    {"name": "__uidac", "domain": ".www.dienmayxanh.com", "value": "016a4343cfc80706dd0a2cd05d400c31"},
    {"name": "__iid", "domain": "www.dienmayxanh.com", "value": ""},
    {"name": "__IP", "domain": "www.dienmayxanh.com", "value": "1906387520"},
    {"name": "__R", "domain": "www.dienmayxanh.com", "value": "3"},
    {"name": "__RC", "domain": "www.dienmayxanh.com", "value": "5"},
    {"name": "__su", "domain": "www.dienmayxanh.com", "value": "0"},
    {"name": "__tb", "domain": "www.dienmayxanh.com", "value": "0"},
    {"name": "__uif", "domain": "www.dienmayxanh.com",
     "value": "__uid%3A7105658824095331495%7C__ui%3A-1%7C__create%3A1780565884"},
    {"name": "_customerIdRecommend", "domain": "www.dienmayxanh.com", "value": "e30cda076bfe1ee5"},
    {"name": "chatmode", "domain": "www.dienmayxanh.com", "value": "1"},
    {"name": "popup_banner_home", "domain": "www.dienmayxanh.com", "value": "popup_banner_H_1days"},
    {"name": "SvID", "domain": "www.dienmayxanh.com", "value": "new26124|aof8d|aof8d"},
]

# ---------------------------------------------------------------------------
# UI styling
# ---------------------------------------------------------------------------
CUSTOM_CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
:root {
    --dmx-accent: #e30613;
    --dmx-accent-soft: #fff1f1;
    --dmx-accent-2: #ff7a45;
}
html, body, [class*="css"] { font-family: 'Inter', -apple-system, 'Segoe UI', sans-serif; }
.dmx-hero {
    position: relative;
    overflow: hidden;
    background: linear-gradient(135deg, #ff7a45 0%, #e30613 55%, #8f0210 100%);
    padding: 30px 34px;
    border-radius: 20px;
    color: #fff;
    margin-bottom: 16px;
    box-shadow: 0 14px 34px rgba(227,6,19,0.30);
}
.dmx-hero::after {
    content: "";
    position: absolute; top: -60px; right: -60px;
    width: 220px; height: 220px; border-radius: 50%;
    background: radial-gradient(circle, rgba(255,255,255,0.16) 0%, rgba(255,255,255,0) 70%);
}
.dmx-hero h1 { margin: 0 0 6px 0; font-size: 26px; font-weight: 800; letter-spacing: -0.02em; }
.dmx-hero p { margin: 0; opacity: 0.95; font-size: 14px; line-height: 1.55; max-width: 720px; }
.dmx-chips { display: flex; flex-wrap: wrap; gap: 8px; margin: 14px 0 20px 0; }
.dmx-chip {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 6px 14px; border-radius: 999px; font-size: 12.5px; font-weight: 600;
    background: var(--dmx-accent-soft); color: var(--dmx-accent);
    border: 1px solid rgba(227,6,19,0.12);
}
.dmx-badge {
    display: inline-block; padding: 2px 10px; border-radius: 999px;
    font-size: 11px; font-weight: 700; background: var(--dmx-accent-soft);
    color: var(--dmx-accent); margin-left: 8px; vertical-align: middle;
}
.dmx-cat-header {
    display: flex; align-items: center; gap: 8px;
    padding: 12px 18px; border-radius: 14px;
    background: linear-gradient(90deg, var(--dmx-accent-soft) 0%, rgba(255,255,255,0) 100%);
    border-left: 4px solid var(--dmx-accent);
    margin: 24px 0 14px 0;
}
.dmx-cat-header h3 { margin: 0; font-size: 16.5px; font-weight: 700; color: #b5030f; }
.dmx-cat-icon { font-size: 20px; }
.dmx-card {
    border: 1px solid rgba(120,120,120,0.16);
    border-radius: 16px;
    padding: 14px;
    margin-bottom: 18px;
    background: rgba(127,127,127,0.02);
    transition: box-shadow .18s ease, transform .18s ease, border-color .18s ease;
}
.dmx-card:hover {
    box-shadow: 0 10px 26px rgba(0,0,0,0.12);
    transform: translateY(-2px);
    border-color: rgba(227,6,19,0.35);
}
.dmx-card-name { font-weight: 700; font-size: 14.5px; margin: 8px 0 3px 0; line-height: 1.35; }
.dmx-card-meta { font-size: 12px; opacity: 0.62; margin-bottom: 8px; }
.dmx-card-meta code {
    background: var(--dmx-accent-soft); color: var(--dmx-accent);
    padding: 1px 6px; border-radius: 6px; font-weight: 600;
}
[data-testid="stImage"] img { border-radius: 12px; }
[data-testid="stMetric"] {
    background: rgba(227,6,19,0.045);
    border: 1px solid rgba(227,6,19,0.10);
    border-radius: 14px;
    padding: 10px 6px 6px 6px;
}
button[kind="primary"] { border-radius: 10px !important; font-weight: 700 !important; }

/* ---- Thanh tab chính (thay st.tabs bằng st.radio để giữ nguyên tab đang
   chọn qua mỗi lần rerun — ví dụ khi bấm nút Lưu trong tab PIM) ---- */
.dmx-tabbar div[data-testid="stRadio"] > label { display: none; }
.dmx-tabbar div[role="radiogroup"] {
    display: flex; flex-wrap: wrap; gap: 6px;
    background: rgba(127,127,127,0.07);
    padding: 6px; border-radius: 14px;
    margin-bottom: 6px;
}
.dmx-tabbar div[role="radiogroup"] label {
    border-radius: 10px !important;
    padding: 9px 18px !important;
    margin: 0 !important;
    font-weight: 600 !important;
    font-size: 14px !important;
    transition: background .15s ease, color .15s ease;
}
.dmx-tabbar div[role="radiogroup"] label:hover { background: rgba(227,6,19,0.09); }
.dmx-tabbar div[role="radiogroup"] label:has(input:checked) {
    background: var(--dmx-accent) !important;
    box-shadow: 0 4px 12px rgba(227,6,19,0.28);
}
.dmx-tabbar div[role="radiogroup"] label:has(input:checked) p { color: #fff !important; font-weight: 700 !important; }
.dmx-tabbar div[role="radiogroup"] label > div:first-child { display: none; }

/* ---- Thanh sub-tab bên trong tab PIM (Chuyển đổi / Cấu hình dữ liệu) — nhỏ
   hơn thanh tab chính, màu trung tính, để phân cấp rõ: đây là điều hướng cấp
   2, không lẫn với 4 tab chính ---- */
.dmx-subtabbar div[role="radiogroup"] {
    background: rgba(127,127,127,0.05);
    border: 1px dashed rgba(127,127,127,0.25);
}
.dmx-subtabbar div[role="radiogroup"] label {
    font-size: 13px !important;
    padding: 7px 14px !important;
}
.dmx-subtabbar div[role="radiogroup"] label:has(input:checked) {
    background: #2b3a55 !important;
    box-shadow: none;
}

/* ---- 2 khu vực Category / CMS-PIM trong tab Chuyển đổi PIM: tách biệt rõ
   ràng bằng tiêu đề màu riêng, dễ phân biệt vùng nào nối-thêm vùng nào đè ---- */
.dmx-zone-title {
    font-weight: 700; font-size: 15px;
    margin: 22px 0 8px 0;
    display: flex; align-items: center; gap: 10px; flex-wrap: wrap;
}
.dmx-zone-cat { color: #1a6b3c; }
.dmx-zone-map { color: var(--dmx-accent); }
.dmx-zone-tag {
    font-size: 11px; font-weight: 700; padding: 3px 10px; border-radius: 999px;
    background: rgba(26,107,60,0.12); color: #1a6b3c;
}
.dmx-zone-tag-alt { background: var(--dmx-accent-soft); color: var(--dmx-accent); }
</style>"""

CATEGORY_ICON_RULES = [
    (("máy lạnh", "điều hòa", "điều hoà"), "❄️"),
    (("tủ lạnh", "tủ đông", "tủ mát"), "🧊"),
    (("máy giặt", "sấy"), "🌀"),
    (("tivi", "tv", "smart tivi"), "📺"),
    (("quạt",), "🌬️"),
    (("bếp", "lò vi sóng", "lò nướng"), "🍳"),
    (("máy lọc nước", "lọc nước"), "💧"),
    (("máy lọc không khí", "lọc không khí"), "🌫️"),
    (("bình nóng lạnh", "nước nóng"), "🚿"),
    (("loa", "âm thanh"), "🔊"),
    (("điện thoại", "laptop", "máy tính"), "💻"),
]


def category_icon(cate_name: str) -> str:
    low = (cate_name or "").lower()
    for keywords, icon in CATEGORY_ICON_RULES:
        if any(k in low for k in keywords):
            return icon
    return "📂"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7",
}


def build_session(cookie_file) -> requests.Session:
    s = requests.Session()
    s.headers.update(HEADERS)
    cookies_to_set = DEFAULT_COOKIES
    if cookie_file is not None:
        try:
            cookie_file.seek(0)
            uploaded = json.load(cookie_file)
            if isinstance(uploaded, list) and uploaded:
                cookies_to_set = uploaded
        except Exception as e:  # noqa: BLE001
            st.warning(f"Không đọc được file cookie vừa tải lên, dùng cookie mặc định thay thế. Lỗi: {e}")
    for c in cookies_to_set:
        name = c.get("name")
        value = c.get("value")
        domain = (c.get("domain") or "").lstrip(".") or "www.dienmayxanh.com"
        if name and value is not None:
            s.cookies.set(name, value, domain=domain)
    return s


def strip_size_suffix(url: str) -> str:
    if not url:
        return url
    return re.sub(r"-\d+x\d+(?=\.\w+(\?.*)?$)", "", url)


def safe_folder_name(name: str, fallback: str) -> str:
    name = (name or fallback or "san-pham").strip()
    name = re.sub(r'[\\/:*?"<>|]+', "-", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:100] if name else fallback


ACCENT_FILL = PatternFill(start_color="FFE30613", end_color="FFE30613", fill_type="solid")
SOFT_FILL = PatternFill(start_color="FFFFF1F1", end_color="FFFFF1F1", fill_type="solid")
WHITE_BOLD = Font(bold=True, color="FFFFFFFF", size=13)
LABEL_BOLD = Font(bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def build_specs_workbook(items) -> bytes:
    """1 sản phẩm = 1 khối: tiêu đề + Mã SP/URL + bảng Thông số/Giá trị — trình bày dọc,
    giống hệt cách hiển thị 'Thông số | Giá trị' trên web, dễ đọc hơn nhiều so với việc
    nhồi hàng chục cột thông số ngang trên 1 dòng."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Thong so ky thuat"
    ws.column_dimensions[get_column_letter(1)].width = 34
    ws.column_dimensions[get_column_letter(2)].width = 85

    row = 1
    for r in items:
        ws.cell(row=row, column=1, value=r["name"] or r["input"]).font = WHITE_BOLD
        ws.cell(row=row, column=1).fill = ACCENT_FILL
        ws.cell(row=row, column=2).fill = ACCENT_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

        for label, value in (
            ("Mã sản phẩm", r.get("product_id") or r["input"]),
            ("URL sản phẩm", r["url"]),
            ("Category", r.get("cate_name") or ""),
            ("Số ảnh gallery", len(r["gallery"])),
        ):
            ws.cell(row=row, column=1, value=label).font = LABEL_BOLD
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Thông số").font = LABEL_BOLD
        ws.cell(row=row, column=2, value="Giá trị").font = LABEL_BOLD
        ws.cell(row=row, column=1).fill = SOFT_FILL
        ws.cell(row=row, column=2).fill = SOFT_FILL
        row += 1

        if r["specs"]:
            for k, v in r["specs"].items():
                ws.cell(row=row, column=1, value=k).alignment = WRAP_TOP
                ws.cell(row=row, column=2, value=v).alignment = WRAP_TOP
                row += 1
        else:
            ws.cell(row=row, column=1, value="(không lấy được thông số)")
            row += 1

        row += 2  # dòng trống ngăn cách giữa các sản phẩm

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_images_workbook(items) -> bytes:
    """Danh sách link ảnh riêng biệt — mỗi dòng 1 ảnh, tách hẳn khỏi bảng thông số để
    không bị dồn chung thành 1 bảng ngang rối mắt."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Link hinh anh"
    headers = ["STT", "Mã SP", "Tên sản phẩm", "STT ảnh", "Link ảnh"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = ACCENT_FILL
    stt = 1
    for r in items:
        for gi, url in enumerate(r["gallery"], 1):
            ws.append([stt, r.get("product_id") or r["input"], r["name"], gi, url])
            stt += 1
    for col, w in zip(range(1, 6), [6, 14, 42, 8, 95]):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CMS (web) -> PIM conversion — reads the "CẤU HÌNH CATEGORY" and
# "THUỘC TÍNH CMS/PIM" tabs LIVE from the shared Google Sheet (kept up to
# date by the team) instead of a bundled snapshot, since that sheet is
# edited continuously. Output values are raw TEXT (no more DATA PIM /
# option-code lookup) per current process.
# ---------------------------------------------------------------------------
PIM_SHEET_ID = "1f2rFnEnVljtNqoG6BCGea7EBuf-OA030HwZURk4ZmY0"
PIM_GID_CATEGORY_CONFIG = "785786646"   # tab: CẤU HÌNH CATEGORY
PIM_GID_CMSPIM_MAPPING = "1125732962"   # tab: THUỘC TÍNH CMS/PIM
PIM_MULTI_JOIN = "|"

# ---------------------------------------------------------------------------
# Lưu trữ lâu dài cấu hình bổ sung (nhập ngay trong tool) lên GitHub, để
# không bị mất khi refresh trang / app redeploy. Đây là 2 kho dữ liệu TÁCH
# BIỆT nhau, đúng theo cách hai loại cấu hình được cập nhật:
#   - category_overrides.json: mỗi lần lưu sẽ NỐI THÊM/mở rộng cột cho từng
#     category (không xoá dữ liệu cũ), giống hệt cách merge_local_category_
#     config() hoạt động khi convert.
#   - cmspim_overrides.json: mỗi lần lưu sẽ ĐÈ lên giá trị cũ của đúng cặp
#     (cate_id, tên thuộc tính) đó — cập nhật thuộc tính nào thì thuộc tính
#     đó lấy giá trị mới nhất, các thuộc tính khác không đổi.
# Đọc (raw.githubusercontent) không cần token vì repo public; ghi (GitHub
# Contents API) cần một Personal Access Token đặt trong Streamlit Secrets
# hoặc biến môi trường với tên GITHUB_TOKEN (quyền "repo" hoặc "contents:write").
# ---------------------------------------------------------------------------
GITHUB_REPO = "ducd64944-ux/dienmayxanh-scraper"
GITHUB_BRANCH = "main"
GH_CATEGORY_OVERRIDE_PATH = "data/category_overrides.json"
GH_MAPPING_OVERRIDE_PATH = "data/cmspim_overrides.json"


def _github_token() -> str:
    try:
        tok = st.secrets.get("GITHUB_TOKEN")  # type: ignore[union-attr]
        if tok:
            return tok
    except Exception:  # noqa: BLE001 - st.secrets raises if no secrets.toml at all
        pass
    return os.environ.get("GITHUB_TOKEN", "")


@st.cache_data(ttl=60, show_spinner=False)
def load_github_json(path: str):
    """Đọc file JSON cấu hình bổ sung đã lưu trên GitHub (public raw, không cần
    token). Trả về [] nếu file chưa tồn tại hoặc lỗi mạng."""
    url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{path}"
    try:
        resp = requests.get(url, timeout=10)
        if resp.status_code == 404:
            return []
        resp.raise_for_status()
        return resp.json() or []
    except Exception:  # noqa: BLE001
        return []


def save_github_json(path: str, data, commit_message: str):
    """Ghi đè file JSON tại `path` trên GitHub qua Contents API (tạo mới nếu
    chưa có, cập nhật đúng sha nếu đã có). Trả về (True, msg) / (False, lỗi)."""
    token = _github_token()
    if not token:
        return False, (
            "Chưa cấu hình GITHUB_TOKEN — vào Streamlit Cloud > app > Settings > "
            "Secrets, thêm dòng GITHUB_TOKEN = \"...\" (Personal Access Token có "
            "quyền ghi vào repo) rồi thử lại."
        )
    api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"
    headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github+json"}
    sha = None
    try:
        r = requests.get(api_url, headers=headers, params={"ref": GITHUB_BRANCH}, timeout=15)
        if r.status_code == 200:
            sha = r.json().get("sha")
    except requests.exceptions.RequestException as e:
        return False, f"Không kết nối được GitHub: {e}"
    content_b64 = base64.b64encode(
        json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    ).decode("ascii")
    payload = {"message": commit_message, "content": content_b64, "branch": GITHUB_BRANCH}
    if sha:
        payload["sha"] = sha
    try:
        r = requests.put(api_url, headers=headers, json=payload, timeout=20)
    except requests.exceptions.RequestException as e:
        return False, f"Không kết nối được GitHub: {e}"
    if r.status_code in (200, 201):
        return True, "Đã lưu lên GitHub."
    return False, f"Lỗi GitHub API ({r.status_code}): {r.text[:300]}"


def build_category_override_payload(local_cat_rows):
    """Chuẩn hoá bảng category đang sửa trong tool thành payload để lưu lên
    GitHub. LƯU Ý: bảng này đã được tự nạp sẵn từ GitHub lúc mở tool (xem chỗ
    seed `local_cat_data` trong UI), nên bản thân bảng đang hiển thị chính là
    toàn bộ dữ liệu hiện có — KHÔNG gộp lại với bản trên GitHub ở bước lưu, để
    nếu người dùng xoá 1 dòng (hoặc xoá bớt mã trong ô "Cột PIM") trong bảng
    rồi bấm Lưu, dòng/mã đó thực sự biến mất trên GitHub thay vì bị merge-nối-
    lại. Dòng nào trùng cate_id thì lấy dòng cuối cùng trong bảng."""
    out = {}
    for row in local_cat_rows or []:
        cid = str(row.get("cate_id") or "").strip()
        if not cid:
            continue
        out[cid] = {
            "cate_id": cid,
            "cate_name": str(row.get("cate_name") or "").strip(),
            "columns_text": str(row.get("columns_text") or "").strip(),
        }
    return list(out.values())


def build_mapping_override_payload(local_map_rows):
    """Chuẩn hoá bảng mapping đang sửa trong tool thành payload để lưu lên
    GitHub. Cùng lý do như build_category_override_payload: không gộp lại với
    bản trên GitHub — bảng hiện tại (đã tự nạp từ GitHub lúc mở tool) là toàn
    bộ dữ liệu, xoá dòng nào trong bảng rồi Lưu thì dòng đó mất trên GitHub
    luôn. Dòng nào trùng (cate_id, tên thuộc tính) thì lấy dòng cuối trong bảng."""
    out = {}
    for row in local_map_rows or []:
        cid = str(row.get("cate_id") or "").strip()
        attr = str(row.get("cms_attr_name") or "").strip()
        if not (cid and attr):
            continue
        out[(cid, attr)] = {
            "cate_id": cid,
            "cms_attr_name": attr,
            "pim_code": str(row.get("pim_code") or "").strip(),
        }
    return list(out.values())


def cross_check_conversion(cat_config, mapping_by_cate, sheets, max_workers=8):
    """Đối chiếu đa luồng (song song theo từng category) ngay sau khi convert
    xong, để bắt các trường hợp khả nghi trước khi người dùng tải file xuống:
      (1) mã PIM có trong THUỘC TÍNH CMS/PIM nhưng KHÔNG có trong danh sách cột
          đã cấu hình cho category đó (map/nhập sai mã, hoặc cấu hình thiếu cột)
      (2) một ô bị gộp từ nhiều giá trị khác nhau bất thường (>=3 giá trị khác
          nhau trong 1 ô) — thường là dấu hiệu nhiều thuộc tính CMS khác nhau
          vô tình map chung vào 1 mã PIM.
      (3) category có sản phẩm nhưng TOÀN BỘ cột TSKT đều trống ở mọi dòng —
          khả năng cao mapping CMS/PIM của cả category đó bị sai/thiếu hoàn
          toàn (vd cate_id giữa 2 sheet không khớp), cần kiểm tra ngay thay vì
          chỉ thấy rải rác từng thuộc tính chưa map.
    Trả về list cảnh báo [{"loai": str, "chi_tiet": str}], rỗng nếu sạch."""
    if not sheets:
        return []

    def check_one(cate_key):
        found = []
        info = sheets[cate_key]
        cfg = cat_config.get(cate_key)
        valid_codes = (
            {c for c, _ in cfg["columns"]} if cfg else {c for c, _ in info["columns"]}
        ) | IDENTITY_CODES
        attr_map = mapping_by_cate.get(cate_key, {})
        for attr, code in attr_map.items():
            if code not in valid_codes:
                found.append({
                    "loai": "Mã PIM không khớp cột đã cấu hình",
                    "chi_tiet": (
                        f"[{info['label']}] thuộc tính '{attr}' → mã '{code}' không có trong "
                        f"danh sách cột đã cấu hình cho category này."
                    ),
                })
        for row in info["rows"]:
            for code, _ in info["columns"]:
                val = row.get(code, "")
                if val and val.count(PIM_MULTI_JOIN) >= 2:
                    parts = [p for p in val.split(PIM_MULTI_JOIN) if p]
                    if len(set(parts)) >= 3:
                        found.append({
                            "loai": "Nhiều giá trị khác nhau bị gộp vào 1 ô",
                            "chi_tiet": (
                                f"[{info['label']}] SP '{row.get('_product_name', '')}', "
                                f"cột '{code}': {len(set(parts))} giá trị khác nhau — {val[:150]}"
                            ),
                        })

        if info["rows"] and info["columns"]:
            all_blank = all(
                not any(str(row.get(code, "")).strip() for code, _ in info["columns"])
                for row in info["rows"]
            )
            if all_blank:
                found.append({
                    "loai": "Category có sản phẩm nhưng KHÔNG map được thuộc tính nào",
                    "chi_tiet": (
                        f"[{info['label']}] {len(info['rows'])} sản phẩm nhưng toàn bộ cột TSKT đều "
                        f"trống — khả năng cao mapping CMS/PIM của category id {cate_key} bị sai/"
                        f"thiếu hoàn toàn, kiểm tra lại sheet THUỘC TÍNH CMS/PIM cho category này."
                    ),
                })
        return found

    warnings = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futs = {ex.submit(check_one, k): k for k in sheets}
        for f in as_completed(futs):
            try:
                warnings.extend(f.result())
            except Exception as e:  # noqa: BLE001
                warnings.append({"loai": "Lỗi khi đối chiếu", "chi_tiet": f"{futs[f]}: {e}"})
    return warnings


def _pim_csv_url(gid: str) -> str:
    return (
        f"https://docs.google.com/spreadsheets/d/{PIM_SHEET_ID}/gviz/tq"
        f"?tqx=out:csv&gid={gid}"
    )


def _norm_attr(s) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s.lower()


@st.cache_data(ttl=600, show_spinner=False)
def load_category_config():
    """Return {cate_id_str: {"name": str, "columns": [(code, viet_name), ...]}}."""
    resp = requests.get(_pim_csv_url(PIM_GID_CATEGORY_CONFIG), timeout=20)
    resp.raise_for_status()
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    rows = rows[1:]  # drop header row
    config = {}
    i = 0
    while i + 1 < len(rows):
        code_row, name_row = rows[i], rows[i + 1]
        cate_id = (code_row[0] if len(code_row) > 0 else "").strip()
        cate_name = (code_row[1] if len(code_row) > 1 else "").strip()
        if cate_id:
            cols = []
            for c_idx in range(2, max(len(code_row), len(name_row))):
                code = code_row[c_idx].strip() if c_idx < len(code_row) else ""
                vname = name_row[c_idx].strip() if c_idx < len(name_row) else ""
                if not code:
                    break
                cols.append((code, vname or code))
            config[cate_id] = {"name": cate_name, "columns": cols}
        i += 2
    return config


@st.cache_data(ttl=600, show_spinner=False)
def load_cmspim_mapping():
    """Return:
       by_cate: {cate_id_str: {norm_attr_name: pim_code}}
       cate_name_to_id: {norm(cate_name): cate_id_str}   (fallback lookup)
       cate_id_to_name: {cate_id_str: cate_name}
    Robust to both the legacy header (MÃ THUỘC TÍNH PIM / TÊN THUỘC TÍNH PIM)
    and the newer one (MÃ MASTER / TÊN MASTER), and skips "▶ NHÓM: ..."
    section-divider rows and rows with no target code yet (still unmapped).
    """
    resp = requests.get(_pim_csv_url(PIM_GID_CMSPIM_MAPPING), timeout=20)
    resp.raise_for_status()
    rows = list(csv.reader(io.StringIO(resp.content.decode("utf-8-sig"))))
    if not rows:
        return {}, {}, {}
    header = [h.strip() for h in rows[0]]

    def find_col(*names):
        for name in names:
            if name in header:
                return header.index(name)
        return None

    idx_cate_id = find_col("MÃ NGÀNH HÀNG CMS")
    idx_cate_name = find_col("TÊN NGÀNH HÀNG CMS")
    idx_attr_name = find_col("TÊN THUỘC TÍNH TSKT")
    idx_code = find_col("MÃ MASTER", "MÃ THUỘC TÍNH PIM")
    if None in (idx_cate_id, idx_cate_name, idx_attr_name, idx_code):
        return {}, {}, {}

    by_cate = defaultdict(dict)
    cate_name_to_id = {}
    cate_id_to_name = {}
    for r in rows[1:]:
        if len(r) <= max(idx_cate_id, idx_cate_name, idx_attr_name, idx_code):
            continue
        cate_id = r[idx_cate_id].strip()
        cate_name = r[idx_cate_name].strip()
        attr_name = r[idx_attr_name].strip()
        code = r[idx_code].strip()
        if not cate_id:
            continue
        cate_id_to_name.setdefault(cate_id, cate_name)
        if cate_name:
            cate_name_to_id.setdefault(_norm_attr(cate_name), cate_id)
        # skip "▶ NHÓM: ..." section-divider rows and not-yet-mapped attrs
        if attr_name.startswith("▶") or "NHÓM" in attr_name.upper():
            continue
        if not code or not attr_name:
            continue
        by_cate[cate_id].setdefault(_norm_attr(attr_name), code)
    return dict(by_cate), cate_name_to_id, cate_id_to_name


def _parse_columns_text(columns_text: str):
    """Parse a "code1:Tên hiển thị 1; code2:Tên hiển thị 2" string (as typed by
    the user in the local override table) into [(code, viet_name), ...]."""
    cols = []
    for part in (columns_text or "").split(";"):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            code, vname = part.split(":", 1)
        else:
            code, vname = part, part
        code = code.strip()
        vname = vname.strip()
        if code:
            cols.append((code, vname or code))
    return cols


def merge_local_category_config(cat_config, local_rows, fallback_names=None):
    """Overlay locally-entered category rows (list of dicts with cate_id,
    cate_name, columns_text) on top of the live-Google-Sheet cat_config. Local
    rows add/extend columns for a category; they never remove existing ones.
    If a row leaves cate_name blank, it's auto-inferred from `fallback_names`
    (a {cate_id: cate_name} lookup — typically built from the site's own
    breadcrumb name, seen while scraping, and/or the THUỘC TÍNH CMS/PIM sheet)
    so the user only has to type/upload the cate_id."""
    fallback_names = fallback_names or {}
    cat_config = {k: {"name": v["name"], "columns": list(v["columns"])} for k, v in cat_config.items()}
    for row in local_rows or []:
        cate_id = str(row.get("cate_id") or "").strip()
        if not cate_id:
            continue
        new_cols = _parse_columns_text(row.get("columns_text"))
        cate_name = str(row.get("cate_name") or "").strip() or fallback_names.get(cate_id, "")
        if cate_id in cat_config:
            existing_codes = {c for c, _ in cat_config[cate_id]["columns"]}
            for c, v in new_cols:
                if c not in existing_codes:
                    cat_config[cate_id]["columns"].append((c, v))
                    existing_codes.add(c)
            if cate_name:
                cat_config[cate_id]["name"] = cate_name
        elif new_cols or cate_name:
            cat_config[cate_id] = {"name": cate_name or cate_id, "columns": new_cols}
    return cat_config


def merge_local_mapping(mapping_by_cate, local_rows):
    """Overlay locally-entered CMS/PIM attribute mappings (list of dicts with
    cate_id, cms_attr_name, pim_code) on top of the live-Google-Sheet mapping.
    Local rows take precedence over the sheet for the same (cate_id, attr)."""
    mapping_by_cate = {k: dict(v) for k, v in mapping_by_cate.items()}
    for row in local_rows or []:
        cate_id = str(row.get("cate_id") or "").strip()
        attr = str(row.get("cms_attr_name") or "").strip()
        code = str(row.get("pim_code") or "").strip()
        if not (cate_id and attr and code):
            continue
        mapping_by_cate.setdefault(cate_id, {})
        mapping_by_cate[cate_id][_norm_attr(attr)] = code
    return mapping_by_cate


def read_override_excel(file, expected_cols):
    """Read a user-supplied xlsx into a DataFrame with exactly `expected_cols`
    columns (in that order). Matches by header name (case/whitespace-insensitive)
    when possible; otherwise falls back to positional (first N columns), so a
    quickly-thrown-together spreadsheet without exact headers still works.
    Fully-empty rows are dropped, everything else is coerced to text."""
    df = pd.read_excel(file)
    df.columns = [str(c).strip() for c in df.columns]
    lower_map = {c.lower(): c for c in df.columns}
    col_map = {}
    for col in expected_cols:
        if col in df.columns:
            col_map[col] = col
        elif col.lower() in lower_map:
            col_map[col] = lower_map[col.lower()]
    if len(col_map) == len(expected_cols):
        df = df[[col_map[c] for c in expected_cols]]
        df.columns = expected_cols
    else:
        df = df.iloc[:, : len(expected_cols)]
        df.columns = expected_cols[: len(df.columns)]
        for c in expected_cols[len(df.columns):]:
            df[c] = ""
        df = df[expected_cols]
    df = df.fillna("").astype(str)
    for c in expected_cols:
        df[c] = df[c].str.strip()
    df = df[(df != "").any(axis=1)].reset_index(drop=True)
    return df


def _clean_id_str(v) -> str:
    """Chuẩn hoá 1 giá trị id đọc từ Excel về text sạch — pandas hay đọc cột
    id toàn số thành float khi cột có ô trống xen kẽ (vd 7358 -> 7358.0), nên
    phải cắt đuôi '.0' nếu có, ngoài ra strip khoảng trắng thường."""
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        s = s[:-2]
    return s


def read_category_columns_excel(file):
    """Đọc file Excel category-template ĐÚNG định dạng thật (kiểu 'Export
    Product Template' — vd bong_den_led_tskt.xlsx / may_lam_sua_hat_tskt.xlsx):
    DÒNG 1 là header — cột A = CATEGORY ID, cột B = CATEGORY NAME, từ cột C
    trở đi CHÍNH LÀ CÁC MÃ TSKT làm luôn tên cột (vd `brand_of_tskt_master`).
    MỖI DÒNG DỮ LIỆU phía dưới = 1 CATEGORY: ô nào ở cột mã TSKT có giá trị
    (tên hiển thị tiếng Việt, vd "Thương hiệu của") thì category đó dùng mã
    cột tương ứng; ô trống = category đó không dùng mã đó. 1 file có thể có
    nhiều category = nhiều dòng dữ liệu, dùng chung 1 hàng header mã cột.
    Trả về DataFrame đúng shape local_cat_data (1 dòng / category,
    columns_text = "code1:Tên 1;code2:Tên 2;...")."""
    df = pd.read_excel(file, header=0)
    df.columns = [str(c).strip() for c in df.columns]
    if len(df.columns) < 3:
        return pd.DataFrame(columns=["cate_id", "cate_name", "columns_text"])
    id_col, name_col = df.columns[0], df.columns[1]
    code_cols = list(df.columns[2:])
    out = []
    for _, row in df.iterrows():
        cate_id = _clean_id_str(row[id_col])
        if not cate_id:
            continue
        cate_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        parts = []
        for code in code_cols:
            val = row[code]
            if pd.isna(val):
                continue
            vname = str(val).strip()
            if not vname:
                continue
            parts.append(f"{code}:{vname}")
        out.append({"cate_id": cate_id, "cate_name": cate_name, "columns_text": "; ".join(parts)})
    return pd.DataFrame(out, columns=["cate_id", "cate_name", "columns_text"])


# Các biến thể tên cột đã gặp trong thực tế cho file Excel THUỘC TÍNH CMS/PIM
# (file tổng hợp đối chiếu CMS-PIM thật của người dùng dùng cả 2 kiểu tên cột
# này tuỳ sheet: sheet tổng hợp dùng kiểu viết hoa toàn bộ, sheet theo từng
# category dùng kiểu Title Case khác chút).
CMSPIM_MAPPING_HEADER_ALIASES = {
    "cate_id": ["MÃ NGÀNH HÀNG CMS", "Mã ngành hàng CMS", "MÃ NGÀNH HÀNG", "cate_id"],
    "attr_name": ["TÊN THUỘC TÍNH TSKT", "Tên thuộc tính TSKT", "cms_attr_name"],
    "code": ["MÃ MASTER", "Mã TSKT (MASTER)", "MÃ THUỘC TÍNH PIM", "pim_code"],
}


def _find_col_ci(columns, *names):
    lower_map = {str(c).strip().lower(): c for c in columns}
    for name in names:
        if name in columns:
            return name
        key = name.strip().lower()
        if key in lower_map:
            return lower_map[key]
    return None


def read_cmspim_mapping_excel(file):
    """Đọc file Excel thuộc tính CMS/PIM thật (vd file tổng hợp đối chiếu
    CMS-PIM) — quét TẤT CẢ sheet trong file (không chỉ sheet đầu tiên, vì file
    thật thường có nhiều sheet: 1 sheet tổng hợp + nhiều sheet theo từng
    category), nhận diện đúng theo TÊN CỘT với các biến thể đã gặp trong thực
    tế (xem CMSPIM_MAPPING_HEADER_ALIASES) — không đọc theo vị trí cột nên
    không bị lấy nhầm cột như trước (đây chính là lý do tool từng chỉ đọc
    được rất ít dòng: file 6 cột không khớp tên cột generic nên bị fallback
    đọc nhầm 3 cột đầu theo vị trí). Sheet nào không đủ 3 cột cần thiết theo
    tên bị bỏ qua; dữ liệu hợp lệ từ mọi sheet khớp được sẽ được GỘP lại (dedup
    theo cate_id + tên thuộc tính, dòng sau đè dòng trước). Bỏ qua dòng
    "▶ NHÓM: ..." và dòng chưa có mã MASTER / chưa có tên thuộc tính."""
    try:
        xls = pd.ExcelFile(file)
    except Exception:  # noqa: BLE001
        return read_override_excel(file, ["cate_id", "cms_attr_name", "pim_code"])

    all_rows = []
    for sheet_name in xls.sheet_names:
        try:
            df = xls.parse(sheet_name, header=0)
        except Exception:  # noqa: BLE001
            continue
        df.columns = [str(c).strip() for c in df.columns]
        col_cate = _find_col_ci(df.columns, *CMSPIM_MAPPING_HEADER_ALIASES["cate_id"])
        col_attr = _find_col_ci(df.columns, *CMSPIM_MAPPING_HEADER_ALIASES["attr_name"])
        col_code = _find_col_ci(df.columns, *CMSPIM_MAPPING_HEADER_ALIASES["code"])
        if not (col_cate and col_attr and col_code):
            continue
        sub = df[[col_cate, col_attr, col_code]].copy()
        sub.columns = ["cate_id", "cms_attr_name", "pim_code"]
        sub["cate_id"] = sub["cate_id"].apply(_clean_id_str)
        sub["cms_attr_name"] = sub["cms_attr_name"].apply(lambda v: "" if pd.isna(v) else str(v).strip())
        sub["pim_code"] = sub["pim_code"].apply(lambda v: "" if pd.isna(v) else str(v).strip())
        sub = sub[~sub["cms_attr_name"].str.startswith("▶")]
        sub = sub[~sub["cms_attr_name"].str.upper().str.contains("NHÓM")]
        sub = sub[(sub["cate_id"] != "") & (sub["cms_attr_name"] != "") & (sub["pim_code"] != "")]
        if not sub.empty:
            all_rows.append(sub)

    if not all_rows:
        # Không sheet nào khớp tên cột chuẩn -> fallback về cách cũ (vị trí 3
        # cột đầu) để vẫn dùng được với file tự soạn không đúng tên cột chuẩn.
        return read_override_excel(file, ["cate_id", "cms_attr_name", "pim_code"])

    out = pd.concat(all_rows, ignore_index=True)
    out = out.drop_duplicates(subset=["cate_id", "cms_attr_name"], keep="last").reset_index(drop=True)
    return out


def match_category(r, cat_config, cate_name_to_id):
    """Find the CẤU HÌNH CATEGORY entry for a scraped product: try the scraped
    cate_id directly first, then fall back to matching by normalized category
    name (site CMS ids and PIM-mapping-sheet ids aren't guaranteed to line up)."""
    cate_id = str(r.get("cate_id") or "").strip()
    if cate_id and cate_id in cat_config:
        return cate_id
    cate_name_norm = _norm_attr(r.get("cate_name") or "")
    if cate_name_norm and cate_name_norm in cate_name_to_id:
        mapped_id = cate_name_to_id[cate_name_norm]
        if mapped_id in cat_config:
            return mapped_id
    return None


def convert_results_to_pim(ok_results, cat_config, mapping_by_cate, cate_name_to_id, cate_id_to_name):
    """Build PIM text-only conversion. Returns:
       sheets: {cate_key: {"columns": [...], "rows": [dict], "label": str}}
       unmatched_category: [r, ...]  (product's category not found in CẤU HÌNH CATEGORY)
       unmatched_attrs: [{"cate": str, "product": str, "attr": str, "value": str}, ...]
    """
    sheets = {}
    unmatched_category = []
    unmatched_attrs = []

    for r in ok_results:
        cfg_cate_id = match_category(r, cat_config, cate_name_to_id)
        if not cfg_cate_id:
            unmatched_category.append(r)
            continue

        cfg = cat_config[cfg_cate_id]
        columns = cfg["columns"]
        # CẤU HÌNH CATEGORY and THUỘC TÍNH CMS/PIM are separate tabs — their category
        # ids aren't guaranteed to line up, so fall back to matching by category name.
        attr_map = mapping_by_cate.get(cfg_cate_id)
        if not attr_map:
            alt_id = next(
                (cid for cid, nm in cate_id_to_name.items() if _norm_attr(nm) == _norm_attr(cfg["name"])),
                None,
            )
            attr_map = mapping_by_cate.get(alt_id, {}) if alt_id else {}

        row_values = defaultdict(list)  # code -> [values...] (handles "gộp" — multiple
        # CMS spec rows mapping to the same PIM code get combined here)
        for spec_name, spec_value in r["specs"].items():
            spec_value = (spec_value or "").strip()
            if not spec_value:
                continue
            code = attr_map.get(_norm_attr(spec_name))
            if not code:
                unmatched_attrs.append(
                    {
                        "cate": cfg["name"] or cfg_cate_id,
                        "cate_id": cfg_cate_id,
                        "product": r["name"],
                        "attr": spec_name,
                        "value": spec_value,
                    }
                )
                continue
            if spec_value not in row_values[code]:
                row_values[code].append(spec_value)

        out_row = {}
        out_row["_product_name"] = r["name"]
        out_row["_product_url"] = r["url"]
        out_row["_product_id"] = r.get("product_id") or r["input"]
        for code, _vname in columns:
            out_row[code] = PIM_MULTI_JOIN.join(row_values.get(code, []))

        key = cfg_cate_id
        if key not in sheets:
            sheets[key] = {"columns": columns, "rows": [], "label": cfg["name"] or cfg_cate_id}
        sheets[key]["rows"].append(out_row)

    return sheets, unmatched_category, unmatched_attrs


IDENTITY_COLUMNS = [("model_code", "Mã model"), ("variant_code", "Mã biến thể")]
IDENTITY_CODES = {c for c, _ in IDENTITY_COLUMNS}


def build_pim_workbook(sheets, unmatched_category, unmatched_attrs) -> bytes:
    """Build the IMPORT-format workbook, 1 sheet per category. `model_code` +
    `variant_code` are ALWAYS the first 2 columns of every category sheet
    (labels "Mã model" / "Mã biến thể"), theo đúng yêu cầu — sau đó mới tới các
    cột TSKT do CẤU HÌNH CATEGORY (+ local override table) quy định, theo đúng
    thứ tự đã cấu hình. Nếu category đã tự khai `model_code`/`variant_code`
    trong danh sách cột thì không bị lặp lại — 2 cột đó vẫn chỉ xuất hiện 1 lần,
    ở đầu. Các cột không có dữ liệu map từ web (bao gồm 2 cột định danh này) để
    trống, người dùng tự điền sau. 3 cột tham chiếu (_product_id/_name/_url)
    vẫn được thêm ở cuối để dễ đối chiếu ngược lại sản phẩm.
    """
    wb = Workbook()
    wb.remove(wb.active)

    ref_cols = ["_product_id", "_product_name", "_product_url"]
    ref_names = ["Mã SP (tham chiếu)", "Tên sản phẩm (tham chiếu)", "URL (tham chiếu)"]

    for cate_key, info in sheets.items():
        title = safe_folder_name(info["label"], cate_key)[:31] or cate_key
        n = 1
        base_title = title
        existing = {ws.title for ws in wb.worksheets}
        while title in existing:
            n += 1
            title = f"{base_title[:28]}_{n}"
        ws = wb.create_sheet(title)

        extra_columns = [(c, n) for c, n in info["columns"] if c not in IDENTITY_CODES]
        codes = [c for c, _ in IDENTITY_COLUMNS] + [c for c, _ in extra_columns]
        names = [n for _, n in IDENTITY_COLUMNS] + [n for _, n in extra_columns]
        header1 = codes + ref_cols
        header2 = names + ref_names
        ws.append(header1)
        ws.append(header2)
        for r in (1, 2):
            for c in range(1, len(header1) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = WHITE_BOLD if r == 1 else LABEL_BOLD
                cell.fill = ACCENT_FILL if r == 1 else SOFT_FILL
                cell.alignment = WRAP_TOP

        for row in info["rows"]:
            ws.append([row.get(c, "") for c in header1])

        for i, c in enumerate(header1, 1):
            ws.column_dimensions[get_column_letter(i)].width = 40 if c in ref_cols else 26
        ws.freeze_panes = "A3"

    if unmatched_category:
        ws = wb.create_sheet("Chưa xác định category")
        ws.append(["Mã SP", "Tên sản phẩm", "cate_id (web)", "cate_name (web)", "URL"])
        for r in unmatched_category:
            ws.append([r.get("product_id") or r["input"], r["name"], r.get("cate_id"), r.get("cate_name"), r["url"]])
        ws.freeze_panes = "A2"

    if unmatched_attrs:
        ws = wb.create_sheet("Thuộc tính chưa map")
        ws.append(["Category", "Sản phẩm", "Tên thuộc tính (web)", "Giá trị"])
        for e in unmatched_attrs:
            ws.append([e["cate"], e["product"], e["attr"], e["value"]])
        for i, w in enumerate([22, 40, 30, 40], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def scrape_one(session: requests.Session, entry: str, retries: int = 2):
    entry = entry.strip()
    if not entry:
        return None

    if entry.isdigit():
        url = f"https://www.dienmayxanh.com/sp-{entry}"
    elif entry.startswith("http"):
        url = entry
    else:
        return {"input": entry, "error": "Không nhận diện được định dạng (không phải id số hoặc link http)."}

    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = session.get(url, timeout=15, allow_redirects=True)
            break
        except Exception as e:  # noqa: BLE001
            last_err = e
            time.sleep(0.5)
    else:
        return {"input": entry, "error": f"Lỗi kết nối sau {retries + 1} lần thử: {last_err}"}

    if resp.status_code == 404 or "ERROR 404" in resp.text[:2000]:
        return {
            "input": entry,
            "error": (
                "404 - không resolve được (thường do link rút gọn /sp-{id} cần cookie/trình "
                "duyệt thật mới redirect đúng, hoặc cookie mặc định đã hết hạn). Hãy mở link "
                "này 1 lần trên trình duyệt và dán URL đầy đủ thay cho id, hoặc tải cookies.json "
                "mới ở mục Cookie nâng cao."
            ),
        }
    if resp.status_code != 200:
        return {"input": entry, "error": f"HTTP {resp.status_code}"}

    soup = BeautifulSoup(resp.text, "html.parser")

    h1 = soup.find("h1")
    name = h1.get_text(strip=True) if h1 else None

    # --- Product id (internal) + category id, from <section data-id=".." data-cate-id="..">
    product_id = None
    cate_id = None
    section = soup.select_one("section.detail[data-id]") or soup.select_one("section[data-cate-id]")
    if section:
        product_id = section.get("data-id")
        cate_id = section.get("data-cate-id")

    # --- Category name + slug, from breadcrumb
    cate_name = None
    cate_slug = None
    breadcrumb = soup.select_one("ul.breadcrumb") or soup.select_one(".breadcrumb")
    if breadcrumb:
        links = breadcrumb.select("a[href]")
        candidates = [a for a in links if a.get("href") and a.get("href") != "/"]
        if candidates:
            last = candidates[-1]
            cate_name = last.get_text(strip=True)
            cate_slug = last.get("href")

    # Gallery images: main slider container
    gallery = []
    container = soup.find(id="slider-default")
    if container:
        seen = set()
        for img in container.find_all("img"):
            for attr in ("src", "data-src", "data-thumb"):
                v = img.get(attr)
                if v and "/Products/Images/" in v:
                    base = strip_size_suffix(v)
                    if base not in seen:
                        seen.add(base)
                        gallery.append(base)
    og_img = soup.find("meta", property="og:image")
    if og_img and og_img.get("content"):
        og_url = strip_size_suffix(og_img["content"])
        seen_urls = {strip_size_suffix(u) for u in gallery}
        if og_url not in seen_urls:
            gallery.insert(0, og_url)

    # Specs
    # Each spec row is normally <li><aside><strong>Label:</strong></aside><aside>...value(s)...</aside></li>.
    # Some rows (e.g. "Tiện ích") pack several values as separate <span class="circle"> tags with no
    # punctuation between them — those are joined explicitly with ", ".
    specs = {}
    spec_root = soup.select_one("#tab-2 .specification-item") or soup.select_one(".specification-item")
    if spec_root:
        for box in spec_root.select(".box-specifi"):
            for li in box.select("ul.text-specifi > li"):
                asides = li.find_all("aside", recursive=False)
                if len(asides) >= 2:
                    label = asides[0].get_text(" ", strip=True).rstrip(":").strip()
                    value_container = asides[1]
                    parts = [
                        c.get_text(" ", strip=True)
                        for c in value_container.find_all(["span", "a"])
                        if c.get_text(strip=True)
                    ]
                    value = ", ".join(parts) if len(parts) > 1 else value_container.get_text(" ", strip=True)
                    if label:
                        specs[label] = re.sub(r"\s+", " ", value.strip())
                else:
                    text = li.get_text(" ", strip=True)
                    if ":" in text:
                        label, value = text.split(":", 1)
                        specs[label.strip()] = re.sub(r"\s+", " ", value.strip())

    return {
        "input": entry,
        "url": resp.url,
        "name": name,
        "product_id": product_id,
        "cate_id": cate_id,
        "cate_name": cate_name,
        "cate_slug": cate_slug,
        "gallery": gallery,
        "specs": specs,
        "error": None,
    }


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
st.markdown(
    """
<div class="dmx-hero">
  <h1>📦 DienmayXANH Product Scraper</h1>
  <p>Lấy tên, mã sản phẩm, category, toàn bộ ảnh gallery và thông số kỹ thuật từ dienmayxanh.com —
  tự nhóm theo category, tải song song nhiều luồng, xuất dữ liệu tách riêng theo category chỉ trong vài cú click.</p>
</div>
<div class="dmx-chips">
  <span class="dmx-chip">🍪 Cookie có sẵn</span>
  <span class="dmx-chip">⚡ Tải đa luồng</span>
  <span class="dmx-chip">🗂️ Tự nhóm Category</span>
  <span class="dmx-chip">📁 Xuất riêng từng Category</span>
  <span class="dmx-chip">📦 ZIP ảnh gốc</span>
</div>
""",
    unsafe_allow_html=True,
)

with st.sidebar:
    st.markdown("### ⚙️ Cấu hình")
    ids_text = st.text_area(
        "Danh sách ID hoặc link (mỗi dòng 1 mục)",
        height=180,
        placeholder="370579\n370580\nhttps://www.dienmayxanh.com/may-lanh/...",
    )
    max_workers = st.slider("⚡ Số luồng tải song song", min_value=1, max_value=10, value=5)

    with st.expander("🍪 Cookie nâng cao", expanded=False):
        st.caption(
            "Tool đã có sẵn cookie mặc định để resolve link rút gọn `/sp-{id}` — không cần tải "
            "cookies.json mỗi lần dùng nữa. Nếu cookie mặc định hết hạn (id bắt đầu lỗi 404 hàng "
            "loạt), tải cookies.json mới lên đây để ghi đè."
        )
        cookies_file = st.file_uploader(
            "Tuỳ chọn: tải cookies.json mới (ghi đè cookie mặc định)", type=["json"]
        )

    run = st.button("🚀 Lấy dữ liệu", type="primary", use_container_width=True)

    with st.expander("ℹ️ Hướng dẫn / lưu ý", expanded=False):
        st.markdown(
            """
- Mỗi dòng nhập **1 ID** (vd: `370579`) hoặc **1 link đầy đủ**.
- Với **ID**, tool resolve qua `https://www.dienmayxanh.com/sp-{id}` bằng cookie mặc định đã
  nhúng sẵn. Nếu id nào không resolve được, mở link đó 1 lần trên trình duyệt rồi dán URL đầy
  đủ thay cho id.
- Kết quả được **tự động nhóm theo Category** (từ `data-cate-id` + breadcrumb).
- Mỗi sản phẩm hiển thị **Mã sản phẩm** đi kèm **URL**.
- Có nút **tải ZIP ảnh** — mỗi sản phẩm 1 thư mục con, nhóm theo category.
            """
        )

if run:
    entries = [line for line in ids_text.splitlines() if line.strip()]
    if not entries:
        st.warning("Hãy nhập ít nhất 1 id hoặc link.")
        st.stop()

    session = build_session(cookies_file)
    results = [None] * len(entries)
    progress = st.progress(0.0, text="Đang lấy dữ liệu...")
    done_count = 0

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {executor.submit(scrape_one, session, entry): i for i, entry in enumerate(entries)}
        for future in as_completed(future_map):
            idx = future_map[future]
            try:
                results[idx] = future.result()
            except Exception as exc:  # noqa: BLE001
                results[idx] = {"input": entries[idx], "error": f"Lỗi không xác định: {exc}"}
            done_count += 1
            progress.progress(done_count / len(entries), text=f"Đã xử lý {done_count}/{len(entries)}")

    progress.empty()
    st.session_state["results"] = results
    st.session_state.pop("image_zip", None)
    st.session_state.pop("export_zip", None)
    st.session_state.pop("pim_workbook", None)
    st.session_state.pop("pim_stats", None)

if "results" in st.session_state:
    results = st.session_state["results"]
    ok_results = [r for r in results if r and not r.get("error")]
    err_results = [r for r in results if r and r.get("error")]

    groups = defaultdict(list)
    for r in ok_results:
        key = (r.get("cate_id") or "?", r.get("cate_name") or "Chưa xác định danh mục")
        groups[key].append(r)

    total_imgs = sum(len(r["gallery"]) for r in ok_results)

    with st.container(border=True):
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("📄 Tổng mục nhập", len(results))
        m2.metric("✅ Thành công", len(ok_results))
        m3.metric("❌ Lỗi", len(err_results))
        m4.metric("🗂️ Category", len(groups))
        m5.metric("🖼️ Tổng ảnh", total_imgs)

    if err_results:
        with st.expander(f"⚠️ {len(err_results)} mục lỗi", expanded=False):
            for r in err_results:
                st.write(f"- **{r['input']}**: {r['error']}")

    if ok_results:
        spec_keys = []
        for r in ok_results:
            for k in r["specs"].keys():
                if k not in spec_keys:
                    spec_keys.append(k)

        # Dùng st.radio (dạng nút bấm ngang, style lại bằng CSS bên dưới) thay cho
        # st.tabs — st.tabs KHÔNG nhớ tab đang chọn qua mỗi lần rerun (ví dụ khi
        # bấm "Lưu lên GitHub" trong tab PIM), nên sẽ tự nhảy về tab đầu tiên mỗi
        # lần bấm nút. st.radio thì giữ nguyên lựa chọn qua session_state, nên
        # bấm Lưu (hay bất kỳ nút nào khác) trong tab nào thì vẫn ở nguyên tab đó.
        MAIN_TABS = ["🗂️ Theo Category", "📋 Bảng dữ liệu", "⬇️ Xuất dữ liệu", "🔄 Chuyển đổi PIM"]
        st.markdown('<div class="dmx-tabbar">', unsafe_allow_html=True)
        active_tab = st.radio(
            "Chuyển khu vực xem",
            MAIN_TABS,
            key="dmx_active_tab",
            horizontal=True,
            label_visibility="collapsed",
        )
        st.markdown("</div>", unsafe_allow_html=True)

        # ---- Tab 1: card grid grouped by category, with live search ----
        if active_tab == MAIN_TABS[0]:
            query = st.text_input("🔍 Tìm theo tên sản phẩm / category / mã SP", "")
            q = query.strip().lower()
            shown_any = False
            for (cate_id, cate_name), items in sorted(groups.items(), key=lambda kv: kv[0][1]):
                filtered = [
                    r for r in items
                    if not q
                    or q in (r["name"] or "").lower()
                    or q in cate_name.lower()
                    or q in str(r.get("product_id") or "").lower()
                ]
                if not filtered:
                    continue
                shown_any = True
                st.markdown(
                    f'<div class="dmx-cat-header"><span class="dmx-cat-icon">{category_icon(cate_name)}</span>'
                    f'<h3>{cate_name}'
                    f'<span class="dmx-badge">cate_id {cate_id}</span>'
                    f'<span class="dmx-badge">{len(filtered)} sản phẩm</span></h3></div>',
                    unsafe_allow_html=True,
                )
                cols = st.columns(3)
                for i, r in enumerate(filtered):
                    with cols[i % 3]:
                        st.markdown('<div class="dmx-card">', unsafe_allow_html=True)
                        if r["gallery"]:
                            st.image(r["gallery"][0], use_container_width=True)
                        st.markdown(f'<div class="dmx-card-name">{r["name"]}</div>', unsafe_allow_html=True)
                        st.markdown(
                            f'<div class="dmx-card-meta">Mã SP: <code>{r.get("product_id") or r["input"]}</code>'
                            f' · {len(r["gallery"])} ảnh · {len(r["specs"])} thông số</div>',
                            unsafe_allow_html=True,
                        )
                        st.markdown(f"[🔗 Xem trên dienmayxanh.com]({r['url']})")
                        with st.expander("Xem chi tiết & thông số"):
                            if len(r["gallery"]) > 1:
                                gcols = st.columns(min(len(r["gallery"]), 4))
                                for gi, gurl in enumerate(r["gallery"]):
                                    gcols[gi % len(gcols)].image(gurl, use_container_width=True)
                            if r["specs"]:
                                st.table(pd.DataFrame(r["specs"].items(), columns=["Thông số", "Giá trị"]))
                            else:
                                st.caption("Không lấy được thông số kỹ thuật.")
                        st.markdown("</div>", unsafe_allow_html=True)
            if not shown_any:
                st.info("Không có sản phẩm nào khớp với từ khoá tìm kiếm.")

        # ---- Tab 2: flat data table ----
        elif active_tab == MAIN_TABS[1]:
            rows = []
            for r in ok_results:
                row = {
                    "Mã SP": r.get("product_id") or r["input"],
                    "Tên sản phẩm": r["name"],
                    "URL sản phẩm": r["url"],
                    "Category ID": r.get("cate_id") or "",
                    "Category": r.get("cate_name") or "",
                    "Số ảnh": len(r["gallery"]),
                    **{k: r["specs"].get(k, "") for k in spec_keys},
                }
                rows.append(row)
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True, hide_index=True)

            st.markdown("**📎 Copy nhanh toàn bộ URL sản phẩm**")
            st.code("\n".join(r["url"] for r in ok_results), language=None)

        # ---- Tab 3: exports ----
        elif active_tab == MAIN_TABS[2]:
            st.markdown("**📁 Xuất dữ liệu — tách riêng theo từng Category**")
            st.caption(
                "Mỗi category ra 1 thư mục riêng trong file ZIP, gồm 2 file tách biệt: "
                "**thong_so_ky_thuat.xlsx** (trình bày dọc từng sản phẩm, giống hệt bảng "
                "Thông số / Giá trị trên web — dễ đọc, không dồn hàng chục cột ngang) và "
                "**link_hinh_anh.xlsx** (chỉ danh sách link ảnh, tách riêng khỏi thông số "
                "cho khỏi rối). Ngoài ra có thêm 1 file tổng quan nhanh ở ngoài cùng."
            )
            if st.button("📁 Chuẩn bị file xuất theo category", use_container_width=True):
                export_buf = io.BytesIO()
                used_folders = set()
                with zipfile.ZipFile(export_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                    overview_buf = io.BytesIO()
                    with pd.ExcelWriter(overview_buf, engine="openpyxl") as writer:
                        for (cate_id, cate_name), items in sorted(groups.items(), key=lambda kv: kv[0][1]):
                            sheet_name = safe_folder_name(cate_name, f"cate-{cate_id}")[:31] or f"cate-{cate_id}"
                            overview_rows = [
                                {
                                    "Mã SP": r.get("product_id") or r["input"],
                                    "Tên sản phẩm": r["name"],
                                    "URL": r["url"],
                                    "Số ảnh": len(r["gallery"]),
                                    "Số thông số": len(r["specs"]),
                                }
                                for r in items
                            ]
                            pd.DataFrame(overview_rows).to_excel(writer, index=False, sheet_name=sheet_name)
                    zf.writestr("00_tong_quan_theo_category.xlsx", overview_buf.getvalue())

                    for (cate_id, cate_name), items in sorted(groups.items(), key=lambda kv: kv[0][1]):
                        folder = safe_folder_name(cate_name, f"category-{cate_id}")
                        candidate = folder
                        n = 1
                        while candidate in used_folders:
                            n += 1
                            candidate = f"{folder} ({n})"
                        used_folders.add(candidate)
                        zf.writestr(f"{candidate}/thong_so_ky_thuat.xlsx", build_specs_workbook(items))
                        zf.writestr(f"{candidate}/link_hinh_anh.xlsx", build_images_workbook(items))

                export_buf.seek(0)
                st.session_state["export_zip"] = export_buf.getvalue()
                st.success("Đã chuẩn bị xong, bấm nút bên dưới để tải về.")

            if "export_zip" in st.session_state:
                st.download_button(
                    "⬇️ Tải file xuất (ZIP, tách theo category)",
                    data=st.session_state["export_zip"],
                    file_name="dienmayxanh_export_theo_category.zip",
                    mime="application/zip",
                    use_container_width=True,
                )

            st.download_button(
                "⬇️ Tải JSON (toàn bộ dữ liệu thô, cho dev)",
                data=json.dumps(ok_results, ensure_ascii=False, indent=2),
                file_name="dienmayxanh_products.json",
                mime="application/json",
            )

            st.divider()
            st.markdown("**📦 Tải ảnh gốc về (ZIP file ảnh thật)**")
            st.caption(
                "Tải toàn bộ ảnh gallery của các sản phẩm đã lấy thành 1 file ZIP duy nhất — mỗi "
                "sản phẩm 1 thư mục con đặt theo tên sản phẩm, được nhóm theo category. Việc tải "
                "ảnh chạy ngay trong trình duyệt của bạn (không qua server) vì CDN ảnh của "
                "dienmayxanh.com chặn tải tự động từ server ngoài."
            )

            zip_items = []
            used_names_js = set()
            for (cate_id, cate_name), items in sorted(groups.items(), key=lambda kv: kv[0][1]):
                cate_folder = safe_folder_name(cate_name, f"category-{cate_id}")
                for r in items:
                    if not r["gallery"]:
                        continue
                    base_name = safe_folder_name(r["name"], r.get("product_id") or r["input"])
                    folder_name = f"{base_name} (ID {r.get('product_id') or r['input']})"
                    n = 1
                    candidate = folder_name
                    while candidate in used_names_js:
                        n += 1
                        candidate = f"{folder_name} ({n})"
                    used_names_js.add(candidate)
                    zip_items.append(
                        {"cate": cate_folder, "product": candidate, "images": r["gallery"]}
                    )

            zip_items_json = json.dumps(zip_items, ensure_ascii=False)
            components.html(
                f"""
<div style="font-family:Inter,-apple-system,'Segoe UI',sans-serif;">
  <button id="dmxZipBtn" style="background:#e30613;color:#fff;border:none;padding:11px 20px;
    border-radius:10px;font-weight:700;font-size:14px;cursor:pointer;width:100%;">
    🖼️ Tải ZIP ảnh (chạy trong trình duyệt của bạn)
  </button>
  <div id="dmxZipStatus" style="margin-top:8px;font-size:13px;color:#666;"></div>
</div>
<script src="https://cdnjs.cloudflare.com/ajax/libs/jszip/3.10.1/jszip.min.js"></script>
<script>
const DMX_ITEMS = {zip_items_json};
document.getElementById('dmxZipBtn').addEventListener('click', async () => {{
  const btn = document.getElementById('dmxZipBtn');
  const status = document.getElementById('dmxZipStatus');
  btn.disabled = true;
  btn.style.opacity = '0.6';
  const zip = new JSZip();
  let total = 0, done = 0, failed = 0;
  DMX_ITEMS.forEach(it => total += it.images.length);
  if (total === 0) {{
    status.textContent = 'Không có ảnh nào để tải.';
    btn.disabled = false;
    btn.style.opacity = '1';
    return;
  }}
  for (const it of DMX_ITEMS) {{
    for (let i = 0; i < it.images.length; i++) {{
      const url = it.images[i];
      try {{
        const resp = await fetch(url);
        if (!resp.ok) throw new Error('HTTP ' + resp.status);
        const blob = await resp.blob();
        const extMatch = url.match(/\\.(\\w+)(?:\\?.*)?$/);
        const ext = extMatch ? extMatch[1] : 'jpg';
        const num = String(i + 1).padStart(2, '0');
        zip.file(it.cate + '/' + it.product + '/' + num + '.' + ext, blob);
      }} catch (e) {{
        failed++;
      }}
      done++;
      status.textContent = 'Đang tải ảnh... ' + done + '/' + total + (failed ? ' (' + failed + ' lỗi)' : '');
    }}
  }}
  status.textContent = 'Đang nén file ZIP...';
  const content = await zip.generateAsync({{type: 'blob'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(content);
  a.download = 'dienmayxanh_images.zip';
  document.body.appendChild(a);
  a.click();
  a.remove();
  status.textContent = 'Xong! Đã tải ' + (total - failed) + '/' + total + ' ảnh vào file ZIP.'
    + (failed ? ' (' + failed + ' ảnh lỗi.)' : '');
  btn.disabled = false;
  btn.style.opacity = '1';
}});
</script>
""",
                height=90,
            )

        # ---- Tab 4: convert scraped TSKT -> PIM import (text-only) ----
        elif active_tab == MAIN_TABS[3]:
            st.markdown("**🔄 Chuyển đổi TSKT vừa lấy sang định dạng PIM**")
            st.caption(
                "Dựa trên 2 sheet **CẤU HÌNH CATEGORY** và **THUỘC TÍNH CMS/PIM** lấy trực tiếp "
                "(live) từ Google Sheet — bạn cập nhật bên đó là tool dùng ngay bản mới nhất, "
                "không cần upload lại gì cả. Giá trị xuất ra là **text thô** (không mã hoá số), "
                "nhiều dòng CMS map cùng 1 cột PIM sẽ được gộp và nối bằng dấu `|`. "
                "**2 cột đầu tiên của MỌI sheet category luôn là `model_code` (Mã model) và "
                "`variant_code` (Mã biến thể)** — tự động thêm vào, không cần khai trong cấu hình "
                "category; 2 cột này chưa map được từ web nên để trống, bạn tự điền sau. Sau đó mới "
                "tới các cột TSKT theo đúng cấu hình của từng category trên sheet CẤU HÌNH CATEGORY. "
                "Có thêm 3 cột tham chiếu (Mã SP / Tên / URL) ở cuối mỗi sheet để dễ đối chiếu ngược "
                "lại sản phẩm."
            )

            # -- Local override area: add/edit Category config + CMS/PIM mapping right
            # here in the tool, for categories/attributes not (yet) on the Google Sheet.
            # These rows are merged on top of the live sheet data at conversion time —
            # no need to wait for the Google Sheet to be updated first.
            # Nạp lại cấu hình bổ sung đã lưu trên GitHub từ lần trước (nếu có) —
            # chỉ nạp 1 lần lúc mới vào (session_state chưa có), không ghi đè
            # những gì người dùng đang gõ dở trong phiên hiện tại.
            if "local_cat_data" not in st.session_state:
                persisted_cat = load_github_json(GH_CATEGORY_OVERRIDE_PATH)
                st.session_state.local_cat_data = (
                    pd.DataFrame(persisted_cat, columns=["cate_id", "cate_name", "columns_text"])
                    if persisted_cat
                    else pd.DataFrame(columns=["cate_id", "cate_name", "columns_text"])
                )
            if "local_map_data" not in st.session_state:
                persisted_map = load_github_json(GH_MAPPING_OVERRIDE_PATH)
                st.session_state.local_map_data = (
                    pd.DataFrame(persisted_map, columns=["cate_id", "cms_attr_name", "pim_code"])
                    if persisted_map
                    else pd.DataFrame(columns=["cate_id", "cms_attr_name", "pim_code"])
                )

            PIM_SUBTABS = ["🔄 Chuyển đổi", "🛠️ Cấu hình dữ liệu (Category & CMS/PIM)"]
            st.markdown('<div class="dmx-tabbar dmx-subtabbar">', unsafe_allow_html=True)
            pim_subtab = st.radio(
                "Khu vực trong PIM",
                PIM_SUBTABS,
                key="dmx_pim_subtab",
                horizontal=True,
                label_visibility="collapsed",
            )
            st.markdown("</div>", unsafe_allow_html=True)
            st.caption(
                "🛠️ Cấu hình dữ liệu (Category, thuộc tính CMS/PIM, lưu GitHub) tách RIÊNG khỏi "
                "bước Chuyển đổi — sang bên đó chỉ để CẬP NHẬT cấu hình, không đụng vào dữ liệu "
                "sản phẩm vừa lấy; bấm Chuyển đổi ở tab này mới thực sự chạy convert."
            )

            if pim_subtab == PIM_SUBTABS[1]:
                # -- Vùng hiển thị ALL: tổng quan toàn bộ category đang có cấu hình
                # (Google Sheet + cấu hình bổ sung đã lưu/đang sửa), để biết đã
                # "import"/cấu hình được bao nhiêu category rồi, không phải bấm
                # Chuyển đổi mới xem được.
                try:
                    overview_cat_config = load_category_config()
                except Exception:  # noqa: BLE001
                    overview_cat_config = {}
                try:
                    overview_mapping_by_cate, _ov_n2i, _ov_i2n = load_cmspim_mapping()
                except Exception:  # noqa: BLE001
                    overview_mapping_by_cate = {}
                overview_cat_config = merge_local_category_config(
                    overview_cat_config, st.session_state.local_cat_data.fillna("").to_dict("records")
                )
                overview_mapping_by_cate = merge_local_mapping(
                    overview_mapping_by_cate, st.session_state.local_map_data.fillna("").to_dict("records")
                )
                overview_rows = [
                    {
                        "cate_id": cid,
                        "Tên category": cfg["name"] or "(chưa đặt tên)",
                        "Số cột PIM đã cấu hình": len(cfg["columns"]),
                        "Số thuộc tính CMS đã map": len(overview_mapping_by_cate.get(cid, {})),
                    }
                    for cid, cfg in overview_cat_config.items()
                ]
                overview_rows.sort(key=lambda r: r["Tên category"])
                oc1, oc2, oc3 = st.columns(3)
                oc1.metric("🗂️ Tổng số category đã cấu hình", len(overview_rows))
                oc2.metric(
                    "📐 Category có cột PIM",
                    sum(1 for r in overview_rows if r["Số cột PIM đã cấu hình"] > 0),
                )
                oc3.metric(
                    "🏷️ Category có thuộc tính đã map",
                    sum(1 for r in overview_rows if r["Số thuộc tính CMS đã map"] > 0),
                )
                with st.expander(f"🌐 Xem ALL — toàn bộ {len(overview_rows)} category đã cấu hình", expanded=False):
                    if overview_rows:
                        st.dataframe(pd.DataFrame(overview_rows), use_container_width=True, hide_index=True)
                    else:
                        st.caption("Chưa có category nào được cấu hình (cả trên Google Sheet lẫn cấu hình bổ sung).")

                has_unmatched = bool(
                    st.session_state.get("pim_unmatched_category_list")
                    or st.session_state.get("pim_unmatched_attrs_list")
                )
                with st.expander(
                    "🛠️ Cấu hình bổ sung: thêm/sửa Category & thuộc tính CMS/PIM ngay trong tool",
                    expanded=has_unmatched,
                ):
                    st.caption(
                        "Dùng khi category hoặc thuộc tính chưa có (hoặc chưa đúng) trên Google Sheet — "
                        "nhập tạm ở đây để chuyển đổi ngay, không cần chờ cập nhật sheet. Cấu hình bổ "
                        "sung sẽ được **gộp thêm vào** dữ liệu Google Sheet lúc bấm Chuyển đổi (không xoá "
                        "gì trên sheet). Khi rảnh, bạn có thể copy các dòng này lên Google Sheet để dùng "
                        "chung về sau."
                    )

                    seed_col1, seed_col2 = st.columns(2)
                    if seed_col1.button(
                        "➕ Nạp category chưa xác định (lần chuyển đổi gần nhất)",
                        use_container_width=True,
                        disabled=not st.session_state.get("pim_unmatched_category_list"),
                    ):
                        seed_rows = [
                            {"cate_id": cid, "cate_name": cname, "columns_text": ""}
                            for cid, cname in st.session_state.get("pim_unmatched_category_list", [])
                        ]
                        existing = st.session_state.local_cat_data
                        st.session_state.local_cat_data = (
                            pd.concat([existing, pd.DataFrame(seed_rows)], ignore_index=True)
                            .drop_duplicates(subset=["cate_id"], keep="first")
                        )
                        st.rerun()
                    if seed_col2.button(
                        "➕ Nạp thuộc tính chưa map (lần chuyển đổi gần nhất)",
                        use_container_width=True,
                        disabled=not st.session_state.get("pim_unmatched_attrs_list"),
                    ):
                        seed_rows = [
                            {"cate_id": cid, "cms_attr_name": attr, "pim_code": ""}
                            for cid, attr in st.session_state.get("pim_unmatched_attrs_list", [])
                        ]
                        existing = st.session_state.local_map_data
                        st.session_state.local_map_data = (
                            pd.concat([existing, pd.DataFrame(seed_rows)], ignore_index=True)
                            .drop_duplicates(subset=["cate_id", "cms_attr_name"], keep="first")
                        )
                        st.rerun()

                    st.markdown(
                        '<div class="dmx-zone-title dmx-zone-cat">📁 Vùng 1 — Category '
                        '<span class="dmx-zone-tag">sửa/xoá gì trong bảng, Lưu là áp dụng đúng vậy trên GitHub</span></div>',
                        unsafe_allow_html=True,
                    )
                    with st.container(border=True):
                        st.caption("Mỗi dòng 1 category; cột PIM nhập dạng `mã:Tên hiển thị` cách nhau bởi `;`")
                        st.session_state.local_cat_data = st.data_editor(
                            st.session_state.local_cat_data,
                            num_rows="dynamic",
                            use_container_width=True,
                            key="local_cat_editor",
                            column_config={
                                "cate_id": st.column_config.TextColumn("cate_id (web hoặc CMS)", width="small"),
                                "cate_name": st.column_config.TextColumn(
                                    "Tên category (để trống cũng được, tool tự suy ra từ cate_id)"
                                ),
                                "columns_text": st.column_config.TextColumn(
                                    "Cột PIM (mã:Tên; mã:Tên; ...)", width="large"
                                ),
                            },
                        )
                        cat_xlsx = st.file_uploader(
                            "📂 Hoặc nạp nhanh từ file Excel template thật (kiểu Export Product "
                            "Template — cột A = CATEGORY ID, cột B = CATEGORY NAME, từ cột C trở đi "
                            "là các MÃ TSKT làm tên cột luôn; mỗi dòng dữ liệu = 1 category, ô nào có "
                            "giá trị (tên hiển thị tiếng Việt) thì category đó dùng mã cột đó, ô trống "
                            "là không dùng — 1 file gộp được nhiều category cùng lúc)",
                            type=["xlsx", "xls"],
                            key="local_cat_xlsx_uploader",
                        )
                        if cat_xlsx is not None:
                            try:
                                new_rows = read_category_columns_excel(cat_xlsx)
                                st.session_state.local_cat_data = (
                                    pd.concat([st.session_state.local_cat_data, new_rows], ignore_index=True)
                                    .drop_duplicates(subset=["cate_id"], keep="last")
                                )
                                st.caption(f"✅ Đã nạp {len(new_rows)} category từ file Excel vào bảng trên.")
                            except Exception as e:  # noqa: BLE001
                                st.error(f"Không đọc được file Excel: {e}")

                        if st.button(
                            "☁️ Lưu Category lên GitHub (đúng như bảng đang hiển thị — xoá dòng nào thì mất dòng đó)",
                            use_container_width=True,
                            key="save_cat_github_btn",
                        ):
                            payload = build_category_override_payload(
                                st.session_state.local_cat_data.fillna("").to_dict("records")
                            )
                            ok, msg = save_github_json(
                                GH_CATEGORY_OVERRIDE_PATH, payload, "Cập nhật category overrides từ tool PIM"
                            )
                            if ok:
                                load_github_json.clear()
                                st.success(msg)
                            else:
                                st.error(msg)

                    st.markdown(
                        '<div class="dmx-zone-title dmx-zone-map">🏷️ Vùng 2 — Thuộc tính CMS/PIM '
                        '<span class="dmx-zone-tag dmx-zone-tag-alt">đè giá trị mới lên thuộc tính vừa sửa</span></div>',
                        unsafe_allow_html=True,
                    )
                    with st.container(border=True):
                        st.caption("Map tên thuộc tính lấy được từ web sang mã cột PIM")
                        st.session_state.local_map_data = st.data_editor(
                            st.session_state.local_map_data,
                            num_rows="dynamic",
                            use_container_width=True,
                            key="local_map_editor",
                            column_config={
                                "cate_id": st.column_config.TextColumn("cate_id (web hoặc CMS)", width="small"),
                                "cms_attr_name": st.column_config.TextColumn("Tên thuộc tính (lấy từ web)"),
                                "pim_code": st.column_config.TextColumn("Mã cột PIM"),
                            },
                        )
                        map_xlsx = st.file_uploader(
                            "📂 Hoặc nạp nhanh từ file Excel đối chiếu CMS/PIM thật (nhận đúng các "
                            "cột: MÃ NGÀNH HÀNG CMS / TÊN THUỘC TÍNH TSKT / MÃ MASTER — hoặc "
                            "cate_id/cms_attr_name/pim_code; quét TẤT CẢ sheet trong file, không chỉ "
                            "sheet đầu)",
                            type=["xlsx", "xls"],
                            key="local_map_xlsx_uploader",
                        )
                        if map_xlsx is not None:
                            try:
                                new_rows = read_cmspim_mapping_excel(map_xlsx)
                                st.session_state.local_map_data = (
                                    pd.concat([st.session_state.local_map_data, new_rows], ignore_index=True)
                                    .drop_duplicates(subset=["cate_id", "cms_attr_name"], keep="last")
                                )
                                st.caption(
                                    f"✅ Đã nạp {len(new_rows)} dòng thuộc tính từ file Excel vào bảng trên "
                                    f"(quét toàn bộ sheet trong file)."
                                )
                            except Exception as e:  # noqa: BLE001
                                st.error(f"Không đọc được file Excel: {e}")

                        if st.button(
                            "☁️ Lưu thuộc tính CMS/PIM lên GitHub (ĐÈ giá trị mới lên đúng thuộc tính vừa sửa)",
                            use_container_width=True,
                            key="save_map_github_btn",
                        ):
                            payload = build_mapping_override_payload(
                                st.session_state.local_map_data.fillna("").to_dict("records")
                            )
                            ok, msg = save_github_json(
                                GH_MAPPING_OVERRIDE_PATH, payload, "Cập nhật CMS/PIM mapping overrides từ tool PIM"
                            )
                            if ok:
                                load_github_json.clear()
                                st.success(msg)
                            else:
                                st.error(msg)

                    st.caption(
                        "ℹ️ 2 kho lưu trên GitHub tách biệt hoàn toàn: `category_overrides.json` và "
                        "`cmspim_overrides.json` — sửa bên nào chỉ ảnh hưởng đúng file của bên đó. Bấm "
                        "Lưu vẫn ở nguyên tab \"🔄 Chuyển đổi PIM\", không bị nhảy tab.\n\n"
                        "🗑️ **Muốn xoá dòng sai (hoặc xoá bớt mã cột PIM bị nhầm):** 2 bảng ở trên đã tự "
                        "nạp đúng dữ liệu đang lưu trên GitHub ngay khi mở tool — cứ xoá thẳng trong "
                        "bảng (bôi đen dòng cần xoá rồi bấm phím Delete/Backspace, hoặc icon thùng rác ở "
                        "cuối dòng khi rê chuột vào; với 1 mã cột PIM bị sai trong ô \"Cột PIM\" thì sửa "
                        "trực tiếp chuỗi `mã:Tên; mã:Tên` trong ô đó), rồi bấm nút ☁️ Lưu tương ứng — dòng/"
                        "mã bị xoá sẽ mất trên GitHub luôn, KHÔNG tự động thêm lại."
                    )

                    dl_col, ul_col = st.columns(2)
                    override_payload = json.dumps(
                        {
                            "category": st.session_state.local_cat_data.fillna("").to_dict("records"),
                            "mapping": st.session_state.local_map_data.fillna("").to_dict("records"),
                        },
                        ensure_ascii=False,
                        indent=2,
                    ).encode("utf-8")
                    dl_col.download_button(
                        "💾 Lưu cấu hình bổ sung ra file (dùng lại lần sau)",
                        data=override_payload,
                        file_name="dmx_pim_cau_hinh_bo_sung.json",
                        mime="application/json",
                        use_container_width=True,
                    )
                    override_file = ul_col.file_uploader(
                        "📂 Nạp lại file cấu hình bổ sung đã lưu", type=["json"], key="override_uploader"
                    )
                    if override_file is not None:
                        try:
                            payload = json.loads(override_file.getvalue().decode("utf-8"))
                            st.session_state.local_cat_data = pd.DataFrame(
                                payload.get("category", []),
                                columns=["cate_id", "cate_name", "columns_text"],
                            )
                            st.session_state.local_map_data = pd.DataFrame(
                                payload.get("mapping", []),
                                columns=["cate_id", "cms_attr_name", "pim_code"],
                            )
                            st.success("Đã nạp cấu hình bổ sung từ file. Kéo lên xem lại bảng ở trên.")
                        except Exception as e:  # noqa: BLE001
                            st.error(f"File không đúng định dạng: {e}")

            else:
                if st.button("🔄 Chuyển đổi sang PIM", use_container_width=True, type="primary"):
                    try:
                        with st.spinner("Đang tải cấu hình category + mapping CMS/PIM từ Google Sheet..."):
                            cat_config = load_category_config()
                            mapping_by_cate, cate_name_to_id, cate_id_to_name = load_cmspim_mapping()
                        if not cat_config or not mapping_by_cate:
                            st.error(
                                "Không tải được dữ liệu cấu hình từ Google Sheet (trống hoặc đổi cấu trúc "
                                "cột). Kiểm tra lại sheet rồi thử lại."
                            )
                        else:
                            local_cat_rows = st.session_state.local_cat_data.fillna("").to_dict("records")
                            local_map_rows = st.session_state.local_map_data.fillna("").to_dict("records")
                            # cate_name auto-inferred from cate_id when left blank. Priority: the
                            # CMS/PIM category data the user themselves loaded into the "Cấu hình
                            # bổ sung" table (typed in, or nạp từ file Excel — cate_id + cate_name
                            # đi cùng nhau) is the authoritative source; the live Google Sheet and
                            # the name scraped from the site's breadcrumb are only used as a last
                            # resort when that cate_id hasn't been loaded locally yet.
                            fallback_names = dict(cate_id_to_name)
                            for r in ok_results:
                                cid = str(r.get("cate_id") or "").strip()
                                cname = r.get("cate_name") or ""
                                if cid and cname:
                                    fallback_names[cid] = cname
                            for row in local_cat_rows:
                                cid = str(row.get("cate_id") or "").strip()
                                cname = str(row.get("cate_name") or "").strip()
                                if cid and cname:
                                    fallback_names[cid] = cname  # user's own loaded data wins
                            cat_config = merge_local_category_config(cat_config, local_cat_rows, fallback_names)
                            mapping_by_cate = merge_local_mapping(mapping_by_cate, local_map_rows)

                            sheets, unmatched_cate, unmatched_attrs = convert_results_to_pim(
                                ok_results, cat_config, mapping_by_cate, cate_name_to_id, cate_id_to_name
                            )
                            st.session_state["pim_workbook"] = build_pim_workbook(
                                sheets, unmatched_cate, unmatched_attrs
                            )
                            st.session_state["pim_stats"] = {
                                "matched_products": sum(len(v["rows"]) for v in sheets.values()),
                                "categories": len(sheets),
                                "unmatched_category": len(unmatched_cate),
                                "unmatched_attrs": len(unmatched_attrs),
                            }
                            # keep the raw unmatched lists (deduped) around so the "seed" buttons
                            # above can pre-fill the local override tables on the next run
                            seen_cate = set()
                            cate_list = []
                            for r in unmatched_cate:
                                key = (str(r.get("cate_id") or ""), r.get("cate_name") or "")
                                if key not in seen_cate:
                                    seen_cate.add(key)
                                    cate_list.append(key)
                            st.session_state["pim_unmatched_category_list"] = cate_list

                            seen_attr = set()
                            attr_list = []
                            for e in unmatched_attrs:
                                key = (str(e.get("cate_id") or ""), e.get("attr") or "")
                                if key not in seen_attr:
                                    seen_attr.add(key)
                                    attr_list.append(key)
                            st.session_state["pim_unmatched_attrs_list"] = attr_list

                            # Đối chiếu chéo đa luồng để bắt các trường hợp khả nghi
                            # (mã PIM sai/không khớp cấu hình, gộp nhầm nhiều giá trị)
                            # trước khi người dùng tải file — chạy song song theo category.
                            st.session_state["pim_cross_check_warnings"] = cross_check_conversion(
                                cat_config, mapping_by_cate, sheets
                            )

                            st.success("Đã chuyển đổi xong, xem kết quả bên dưới.")
                    except requests.exceptions.RequestException as e:
                        st.error(f"Không kết nối được tới Google Sheet: {e}")

                if "pim_stats" in st.session_state:
                    stats = st.session_state["pim_stats"]
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("✅ Sản phẩm đã map", stats["matched_products"])
                    c2.metric("🗂️ Category ra file", stats["categories"])
                    c3.metric("❓ Không xác định category", stats["unmatched_category"])
                    c4.metric("⚠️ Thuộc tính chưa map", stats["unmatched_attrs"])
                    if stats["unmatched_category"] or stats["unmatched_attrs"]:
                        st.warning(
                            f"⚠️ Thiếu cấu hình: {stats['unmatched_category']} category chưa xác định "
                            f"được, {stats['unmatched_attrs']} thuộc tính chưa map được mã PIM. Chi tiết "
                            "nằm trong sheet \"Chưa xác định category\" / \"Thuộc tính chưa map\" của file "
                            "tải về, hoặc bấm nút \"➕ Nạp...\" ở khung Cấu hình bổ sung phía trên để điền "
                            "nhanh ngay trong tool rồi chuyển đổi lại."
                        )

                    cross_check_warnings = st.session_state.get("pim_cross_check_warnings") or []
                    if cross_check_warnings:
                        with st.expander(
                            f"🧵 Đối chiếu đa luồng phát hiện {len(cross_check_warnings)} trường hợp khả nghi",
                            expanded=True,
                        ):
                            for w in cross_check_warnings[:100]:
                                st.markdown(f"- **{w['loai']}**: {w['chi_tiet']}")
                            if len(cross_check_warnings) > 100:
                                st.caption(f"... và {len(cross_check_warnings) - 100} trường hợp khác (đã ẩn bớt).")
                    else:
                        st.caption("🧵 Đối chiếu đa luồng: không phát hiện bất thường.")

                    st.download_button(
                        "⬇️ Tải file PIM (xlsx, 1 sheet / category)",
                        data=st.session_state["pim_workbook"],
                        file_name="dienmayxanh_pim_import.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
